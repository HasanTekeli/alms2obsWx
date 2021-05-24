import os
import openpyxl
import pandas as pd
from deps_list import deps_list


# Mods klasörünü oluştur ve düzenlenmiş dosyaları buraya kaydet
def organize_results(folder_path, list_of_deps):
    os.chdir(folder_path)
    files = os.listdir(folder_path)

    try:
        os.mkdir("mods")
        os.mkdir("mods/day")
        os.mkdir("mods/night")
    except FileExistsError:
        pass
    mods_path = folder_path + "/mods/"
    idx = 1
    for f in files:
        if f.endswith('.xlsx'):
            file_path = str(folder_path) + "/{}".format(f)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            dep = ws['C2']
            dep_value = dep.value
            # C2 hücresinde bölüm ismini bulup dosya isimlerini ayarlamak için:
            if dep_value:
                for department in deps_list:
                    if department in dep_value:
                        dep_name = department
                        list_of_deps[idx] = dep_name
                        idx += 1
                    
            

                ws.delete_cols(2, 2)  # B:C arası gereksizleri sil
                ws.delete_cols(3, 6)  # J sil
                ws.delete_rows(1, 1)  # 1.satırı sil

                col_A = []
                col_B = []

                ws.column_dimensions['A'].width = 20  # A sütun genişiliğini azaltma
                for col_cell in ws['A']:  # A sütunundaki hücreleri number formatına getirir.
                    try:
                        col_cell.number_format = '0'
                        col_cell.value = int(col_cell.value)
                        col_A.append(col_cell.value)
                    except ValueError:  # Bu kısma öğrenci numarası ALMS'de görünmeyen öğrenciler için bir uyarı eklenecek.
                        pass

                for col_cell in ws['B']:
                    if col_cell.value != None:
                        try:
                            col_cell.number_format = '0'  # B sütunundaki hücreleri number formatına getirir.
                            col_cell.value = int(
                                float(col_cell.value.replace(",", ".")))  # Sayı decimal işaretini ,'den .'ya çevirir.
                        except AttributeError:
                            pass
                    else:
                        pass
                    col_B.append(col_cell.value)

                all_data = zip(col_A, col_B)  # col_A ve col_B'de topladıklarımızı eşleştir.
                all_data_list = list(all_data)  # üsttekini listeye çevir
                sorted_list = sorted(all_data_list, key=lambda x: x[1], reverse=True)  # Notları büyükten küçüğe sıralar
                df = pd.DataFrame(sorted_list)  # Sıralanmış halini pandas dataframe'e çevirir

                if "(İÖ)" not in dep_value:
                    try:
                        writer = pd.ExcelWriter("mods/day/{}.xlsx".format(dep_name), engine='xlsxwriter')
                        df.to_excel(writer, index=False, sheet_name='Sheet1')
                        wb_alldeps = writer.book
                        ws_alldeps = writer.sheets['Sheet1']
                        fmt = wb_alldeps.add_format({'num_format': '0'})  # Sayı formatını belirle
                        ws_alldeps.set_column('A:B', 20, fmt)  # Gerekli sütunları seç ve sayı formatını uygula
                        writer.save()
                    except FileNotFoundError:
                        pass
                else:
                    try:
                        dep_name = dep_name + " (İÖ)"
                        writer = pd.ExcelWriter("mods/night/{}.xlsx".format(dep_name), engine='xlsxwriter')
                        df.to_excel(writer, index=False, sheet_name='Sheet1')
                        wb_alldeps = writer.book
                        ws_alldeps = writer.sheets['Sheet1']
                        fmt = wb_alldeps.add_format({'num_format': '0'})  # Sayı formatını belirle
                        ws_alldeps.set_column('A:B', 20, fmt)  # Gerekli sütunları seç ve sayı formatını uygula
                        writer.save()
                    except FileNotFoundError:
                        pass
    return mods_path


# Mods klasöründeki gerekli klasörü aç ve panoya kopyala
def create_data(folder_path, id):
    data = []
    chosen_file = folder_path + id + ".xlsx"
    wb = openpyxl.load_workbook(chosen_file)
    ws = wb.active
    copied_cells = ws['A2:B200']
    for a, b in copied_cells:
        if a.value and b.value != -1:
            data.append([a.value, b.value])

    df = pd.DataFrame(data=data, columns=["Num", "Grade"])
    tr_list = df.drop_duplicates(subset=["Num"])  # ALMS'nin zaman zaman verdiği birden fazla kaydı engellemek için
    tr_list.to_clipboard(excel=True, header=False, index=False)